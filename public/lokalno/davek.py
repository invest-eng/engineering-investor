#!/usr/bin/env python3
"""
=============================================================
  Davcni tracker v1.0 -- FIFO kapitalski dobicki (ZDoh-2)
  engineeringinvestor.si/davek
=============================================================

Lokalna razlicica za vse ki ne zelijo spletnega vmesnika.
Vsi podatki ostanejo na tvojem racunalniku.

UPORABA
-------
  python davek.py                            # bere transakcije.csv
  python davek.py --ibkr ibkr.csv            # uvozi iz IBKR CSV
  python davek.py --leto 2024                # filtriraj po letu
  python davek.py --xml --profil profil.json # izvozi tudi eDavki XML
  python davek.py --pomoč                    # vse moznosti

ZAHTEVE
-------
  Python 3.8 ali novejsi (python.org)
  openpyxl  -- samo za Excel izvoz (pip install openpyxl)

Za XML in CSV ni nobenih dodatnih knjiznic.
"""

import csv
import sys
import json
import re
import argparse
from datetime import datetime
from pathlib import Path


# ================================================================
#  DAVCNE STOPNJE  (ZDoh-2, cl. 97)
# ================================================================

TAX_BRACKETS = [
    (7300, 0.00),   # 20 let ali vec  --> 0 %
    (5475, 0.10),   # 15 do 20 let   --> 10 %
    (3650, 0.15),   # 10 do 15 let   --> 15 %
    (1825, 0.20),   #  5 do 10 let   --> 20 %
    (   0, 0.25),   #  0 do  5 let   --> 25 %
]


def get_tax_rate(holding_days: int) -> float:
    for days, rate in TAX_BRACKETS:
        if holding_days >= days:
            return rate
    return 0.25


def days_between(date_a: str, date_b: str) -> int:
    d1 = datetime.strptime(date_a, "%Y-%m-%d").date()
    d2 = datetime.strptime(date_b, "%Y-%m-%d").date()
    return (d2 - d1).days


def r2(n) -> float:
    return round(float(n), 2)


# ================================================================
#  FIFO IZRACUN
# ================================================================

def calc_fifo(trades: list) -> tuple:
    """
    Vrne (realized_gains, unrealized_lots, errors).

    realized_gains: seznam slovarjev z rezultati FIFO parov
    unrealized_lots: seznam preostalih odprtih lotov
    errors: seznam opisov napak (prodaja brez nakupa)
    """
    sorted_trades = sorted(trades, key=lambda t: t["datum"])

    lots = {}        # ticker -> [lot, ...]
    realized = []
    errors = []

    for trade in sorted_trades:
        key = trade["ticker"].upper().strip()
        if not key:
            continue

        rate  = float(trade.get("tecaj_eur") or 1)
        qty   = float(trade["kolicina"])
        price = float(trade["cena"])
        fees  = float(trade.get("provizija") or 0)

        if qty == 0:
            continue

        # Nabavna cena na enoto vkljucno s provizijo
        cost_eur = price * rate + (fees * rate) / qty

        if trade["tip"] == "nakup":
            lots.setdefault(key, []).append({
                "buy_date":         trade["datum"],
                "cost_per_unit_eur": r2(cost_eur),
                "qty_remaining":    qty,
                "isin":             trade.get("isin", ""),
                "asset_type":       trade.get("vrsta", "stock"),
                "name":             trade.get("ime", "") or trade["ticker"],
                "ticker":           trade["ticker"],
            })

        elif trade["tip"] == "prodaja":
            proceeds_eur = price * rate - (fees * rate) / qty
            remaining    = qty

            while remaining > 1e-9:
                if not lots.get(key):
                    errors.append(
                        f"Ni kupne pozicije za {key} na datum {trade['datum']}"
                    )
                    break

                lot     = lots[key][0]
                matched = min(remaining, lot["qty_remaining"])
                hdays   = days_between(lot["buy_date"], trade["datum"])
                rate_t  = get_tax_rate(hdays)
                acq     = r2(lot["cost_per_unit_eur"] * matched)
                disp    = r2(proceeds_eur * matched)
                gain    = r2(disp - acq)

                realized.append({
                    "ticker":        trade["ticker"],
                    "isin":          lot["isin"],
                    "asset_type":    lot["asset_type"],
                    "name":          lot["name"],
                    "buy_date":      lot["buy_date"],
                    "sell_date":     trade["datum"],
                    "quantity":      r2(matched),
                    "acq_value_eur": acq,
                    "disp_value_eur":disp,
                    "gain_eur":      gain,
                    "holding_days":  hdays,
                    "tax_rate":      rate_t,
                    "tax_amount":    r2(gain * rate_t) if gain > 0 else 0,
                })

                lot["qty_remaining"] = r2(lot["qty_remaining"] - matched)
                remaining            = r2(remaining - matched)
                if lot["qty_remaining"] < 1e-9:
                    lots[key].pop(0)

    unrealized = [
        l for lst in lots.values() for l in lst if l["qty_remaining"] > 1e-9
    ]
    return realized, unrealized, errors


# ================================================================
#  BRANJE TRANSAKCIJ -- lastni CSV format
# ================================================================

REQUIRED_COLS = {"datum", "tip", "ticker", "kolicina", "cena"}

def read_transactions_csv(path: Path) -> list:
    """
    Prebere CSV datoteko s transakcijami.

    Obvezni stolpci: datum, tip (nakup/prodaja), ticker, kolicina, cena
    Neobvezni: vrsta, isin, ime, valuta, tecaj_eur, provizija, broker

    Format datuma: YYYY-MM-DD
    Decimalna locila: pika (.) ali vejica (,) -- obe delujeta
    """
    trades = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        cols = set(reader.fieldnames or [])

        missing = REQUIRED_COLS - cols
        if missing:
            print(f"  NAPAKA: CSV nima stolpcev: {', '.join(missing)}")
            print(f"  Najdeni stolpci: {', '.join(cols)}")
            sys.exit(1)

        for i, row in enumerate(reader, 2):  # 2 ker je 1 glava
            if not any(row.values()):
                continue  # preskoči prazne vrstice
            try:
                trades.append({
                    "datum":    row["datum"].strip(),
                    "tip":      row["tip"].strip().lower().replace("buy", "nakup").replace("sell", "prodaja"),
                    "vrsta":    row.get("vrsta", "stock").strip() or "stock",
                    "ticker":   row["ticker"].strip().upper(),
                    "isin":     row.get("isin", "").strip(),
                    "ime":      row.get("ime", "").strip(),
                    "kolicina": float(row["kolicina"].replace(",", ".")),
                    "cena":     float(row["cena"].replace(",", ".")),
                    "valuta":   row.get("valuta", "EUR").strip() or "EUR",
                    "tecaj_eur":float(row["tecaj_eur"].replace(",", ".")) if row.get("tecaj_eur", "").strip() else 1.0,
                    "provizija":float(row["provizija"].replace(",", ".")) if row.get("provizija", "").strip() else 0.0,
                    "broker":   row.get("broker", "").strip(),
                })
            except (ValueError, KeyError) as e:
                print(f"  Opozorilo: vrstica {i} preskocena -- {e}")

    return trades


# ================================================================
#  BRANJE IBKR CSV (Activity Statement)
# ================================================================

def read_ibkr_csv(path: Path) -> tuple:
    """
    Razcleni IBKR Activity Statement CSV.
    Vrne (trades, warnings).
    """
    trades   = []
    warnings = []
    headers  = []

    with open(path, newline="", encoding="utf-8-sig") as f:
        content = f.read()

    for line in content.splitlines():
        cols = _parse_csv_line(line)
        if len(cols) < 3:
            continue

        if cols[0] == "Trades" and cols[1] == "Header":
            headers = cols[2:]
            continue

        if cols[0] != "Trades" or cols[1] != "Data":
            continue

        row = {headers[j]: v.strip() for j, v in enumerate(cols[2:]) if j < len(headers)}

        disc = row.get("DataDiscriminator", "")
        if disc in ("SubTotal", "Total", "Trade", "ClosedLot"):
            continue

        cat = row.get("Asset Category", "")
        if cat == "Forex" or "Option" in cat or "Future" in cat:
            warnings.append(f"Preskoceno: {cat} ({row.get('Symbol', '')})")
            continue
        if cat not in ("Stocks", "Crypto"):
            if cat:
                warnings.append(f"Neznana kategorija: {cat} ({row.get('Symbol', '')})")
            continue

        qty_raw = _f(row.get("Quantity", "0"))
        if qty_raw == 0:
            continue

        date_raw   = row.get("Date/Time") or row.get("TradeDate", "")
        parsed_date = _parse_ibkr_date(date_raw)
        if not parsed_date:
            warnings.append(f"Neveljaven datum: '{date_raw}' za {row.get('Symbol', '')}")
            continue

        currency = row.get("Currency", "USD")
        trades.append({
            "datum":    parsed_date,
            "tip":      "nakup" if qty_raw > 0 else "prodaja",
            "vrsta":    "kripto" if cat == "Crypto" else "delnica",
            "ticker":   (row.get("Symbol") or "").upper(),
            "isin":     row.get("ISIN", ""),
            "ime":      row.get("Description") or row.get("Symbol", ""),
            "kolicina": abs(qty_raw),
            "cena":     abs(_f(row.get("T. Price") or row.get("TradePrice", "0"))),
            "valuta":   currency,
            "tecaj_eur":1.0 if currency == "EUR" else 0.0,
            "provizija":abs(_f(row.get("Comm/Fee") or row.get("IBCommission", "0"))),
            "broker":   "IBKR",
        })

    return trades, warnings


def _parse_ibkr_date(raw: str):
    if not raw:
        return None
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", raw)
    if m:
        return m.group(1)
    m = re.match(r"^(\d{4})(\d{2})(\d{2})", raw)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def _parse_csv_line(line: str) -> list:
    result, current, in_quotes = [], "", False
    for ch in line:
        if ch == '"':
            in_quotes = not in_quotes
        elif ch == "," and not in_quotes:
            result.append(current)
            current = ""
        else:
            current += ch
    result.append(current)
    return result


def _f(s) -> float:
    try:
        return float(str(s).replace(",", "") or "0")
    except (ValueError, TypeError):
        return 0.0


# ================================================================
#  EXCEL IZVOZ
# ================================================================

def export_excel(trades, realized, unrealized, year, out_file: str) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("\n  NAPAKA: openpyxl ni namesecen.")
        print("  Namesti ga z:  pip install openpyxl")
        print("  Podatke sem shranil v CSV (brez Excela).")
        _fallback_csv(trades, realized, unrealized, year, out_file)
        return False

    wb = Workbook()

    # --- List 1: Transakcije ---
    ws = wb.active
    ws.title = "Transakcije"
    _hrow(ws, ["Datum", "Tip", "Vrsta", "Ticker", "ISIN", "Ime",
               "Kolicina", "Cena", "Valuta", "Tecaj EUR", "Provizija", "Broker"])
    t_list = [t for t in trades if not year or t["datum"].startswith(str(year))]
    t_list.sort(key=lambda t: t["datum"])
    for t in t_list:
        ws.append([
            t["datum"],
            "Nakup" if t["tip"] == "nakup" else "Prodaja",
            t.get("vrsta", ""),
            t["ticker"],
            t.get("isin", ""),
            t.get("ime", ""),
            t["kolicina"],
            t["cena"],
            t.get("valuta", "EUR"),
            t.get("tecaj_eur", 1),
            t.get("provizija", 0),
            t.get("broker", ""),
        ])

    # --- List 2: FIFO dobicki ---
    ws2 = wb.create_sheet("FIFO Dobicki")
    _hrow(ws2, ["Ticker", "ISIN", "Ime", "Vrsta", "Datum nakupa", "Datum prodaje",
                "Dni drzanja", "Kolicina", "Nabavna EUR", "Prodajna EUR",
                "Dobicek EUR", "Davek %", "Davek EUR"])
    gains = [g for g in realized if not year or g["sell_date"].startswith(str(year))]
    for g in gains:
        ws2.append([
            g["ticker"], g["isin"], g["name"], g["asset_type"],
            g["buy_date"], g["sell_date"], g["holding_days"], g["quantity"],
            g["acq_value_eur"], g["disp_value_eur"], g["gain_eur"],
            f"{g['tax_rate']*100:.0f}%", g["tax_amount"],
        ])
    if gains:
        ws2.append([])
        tot_gain = sum(g["gain_eur"] for g in gains)
        tot_tax  = sum(g["tax_amount"] for g in gains)
        row = ws2.max_row + 1
        ws2.append(["SKUPAJ", "", "", "", "", "", "", "",
                    "", "", r2(tot_gain), "", r2(tot_tax)])
        for cell in ws2[ws2.max_row]:
            cell.font = Font(bold=True)

    # --- List 3: Odprte pozicije ---
    ws3 = wb.create_sheet("Odprte pozicije")
    _hrow(ws3, ["Ticker", "ISIN", "Ime", "Vrsta", "Datum nakupa",
                "Preostala kolicina", "Nabavna cena/enoto EUR", "Skupaj EUR"])
    for l in unrealized:
        ws3.append([
            l["ticker"], l["isin"], l["name"], l["asset_type"],
            l["buy_date"], l["qty_remaining"], l["cost_per_unit_eur"],
            r2(l["qty_remaining"] * l["cost_per_unit_eur"]),
        ])

    # Sirina stolpcev
    for ws in (wb.active, ws2, ws3):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    wb.save(out_file)
    return True


def _hrow(ws, headers):
    from openpyxl.styles import Font, PatternFill
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="0A1628")


def _fallback_csv(trades, realized, unrealized, year, base):
    """Ce openpyxl ni namesecen, shrani podatke v navadne CSV datoteke."""
    stem = base.replace(".xlsx", "")

    t_file = stem + "-transakcije.csv"
    with open(t_file, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["datum","tip","vrsta","ticker","isin","ime","kolicina",
                    "cena","valuta","tecaj_eur","provizija","broker"])
        for t in trades:
            w.writerow([t["datum"], t["tip"], t.get("vrsta",""), t["ticker"],
                        t.get("isin",""), t.get("ime",""), t["kolicina"],
                        t["cena"], t.get("valuta","EUR"), t.get("tecaj_eur",1),
                        t.get("provizija",0), t.get("broker","")])
    print(f"  Shranjeno: {t_file}")

    g_file = stem + "-fifo.csv"
    gains = [g for g in realized if not year or g["sell_date"].startswith(str(year))]
    with open(g_file, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["ticker","isin","name","buy_date","sell_date","holding_days",
                    "quantity","acq_value_eur","disp_value_eur","gain_eur",
                    "tax_rate","tax_amount"])
        for g in gains:
            w.writerow([g["ticker"],g["isin"],g["name"],g["buy_date"],g["sell_date"],
                        g["holding_days"],g["quantity"],g["acq_value_eur"],
                        g["disp_value_eur"],g["gain_eur"],g["tax_rate"],g["tax_amount"]])
    print(f"  Shranjeno: {g_file}")


# ================================================================
#  EDAVKI XML IZVOZ  (Doh-KDVP-9)
# ================================================================

def export_edavki_xml(realized, profile: dict, year: int, out_file: str) -> bool:
    gains = [g for g in realized if g["sell_date"].startswith(str(year))]
    if not gains:
        print(f"  Ni realiziranih dobickov za leto {year}.")
        return False

    by_ticker = {}
    for g in gains:
        k = g["ticker"].upper()
        if k not in by_ticker:
            by_ticker[k] = {**g, "rows": []}
        by_ticker[k]["rows"].append(g)

    row_id = 1
    items_xml = []
    for item in by_ticker.values():
        is_fund   = item["asset_type"] in ("etf", "fund", "sklad")
        is_crypto = item["asset_type"] in ("crypto", "kripto")

        rows_xml = []
        for g in item["rows"]:
            rows_xml.append(f"""
        <row>
          <ID>{row_id}</ID>
          <Code>{_esc(g['ticker'])}</Code>
          <ISIN>{_esc(g['isin'])}</ISIN>
          <Name>{_esc(g['name'] or g['ticker'])}</Name>
          <IsFond>{'true' if is_fund else 'false'}</IsFond>
          <Purchase>
            <F1>{g['buy_date']}</F1>
            <F2>0</F2>
            <F3>{g['quantity']}</F3>
            <F4>{g['acq_value_eur']:.2f}</F4>
            <F5>0</F5>
            <F11>{g['acq_value_eur']:.2f}</F11>
          </Purchase>
          <Sale>
            <F6>{g['sell_date']}</F6>
            <F7>{g['quantity']}</F7>
            <F9>{g['disp_value_eur']:.2f}</F9>
            <F10>{g['gain_eur']:.2f}</F10>
          </Sale>
        </row>""")
            row_id += 1

        items_xml.append(f"""
      <KDVPItem>
        <InventoryListType>{'PLVPVK' if is_crypto else 'PLVP'}</InventoryListType>
        <Name>{_esc(item['name'] or item['ticker'])}</Name>
        <HasForeignTax>false</HasForeignTax>
        <HasLossTransfer>false</HasLossTransfer>
        <ForeignTransfer>false</ForeignTransfer>
        <TaxDecreaseConf>false</TaxDecreaseConf>
        <Securities>{''.join(rows_xml)}
        </Securities>
      </KDVPItem>""")

    xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Envelope xmlns="http://edavki.durs.si/Documents/Schemas/Doh_KDVP_9.xsd"
          xmlns:edp="http://edavki.durs.si/Documents/Schemas/EDP-Common-1.xsd"
          xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <edp:Header>
    <edp:taxpayer>
      <edp:taxNumber>{_esc(profile.get('taxNumber',''))}</edp:taxNumber>
      <edp:taxpayerType>FO</edp:taxpayerType>
      <edp:name>{_esc(profile.get('name',''))}</edp:name>
      <edp:address1>{_esc(profile.get('address',''))}</edp:address1>
      <edp:city>{_esc(profile.get('city',''))}</edp:city>
      <edp:postNumber>{_esc(profile.get('postCode',''))}</edp:postNumber>
      <edp:birthDate>{_esc(profile.get('birthDate',''))}</edp:birthDate>
    </edp:taxpayer>
    <edp:Workflow>
      <edp:DocumentWorkflowID>O</edp:DocumentWorkflowID>
    </edp:Workflow>
  </edp:Header>
  <edp:AttachmentList/>
  <edp:Signatures/>
  <body>
    <Doh_KDVP>
      <KDVP>
        <DocumentWorkflowID>O</DocumentWorkflowID>
        <Year>{year}</Year>
        <PeriodStart>{year}-01-01</PeriodStart>
        <PeriodEnd>{year}-12-31</PeriodEnd>
        <IsResident>true</IsResident>
        <TelephoneNumber>{_esc(profile.get('phone',''))}</TelephoneNumber>
        <Email>{_esc(profile.get('email',''))}</Email>
      </KDVP>{''.join(items_xml)}
    </Doh_KDVP>
  </body>
</Envelope>"""

    with open(out_file, "w", encoding="utf-8") as f:
        f.write(xml)
    return True


def _esc(s: str) -> str:
    return (str(s or "")
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))


# ================================================================
#  POVZETEK V TERMINALU
# ================================================================

def print_summary(realized, unrealized, errors, year):
    gains = [g for g in realized if not year or g["sell_date"].startswith(str(year))]
    label = f" za leto {year}" if year else ""

    print()
    if errors:
        print("NAPAKE:")
        for e in errors:
            print(f"  ! {e}")
        print()

    if not gains:
        print(f"Ni realiziranih dobickov{label}.")
    else:
        pos    = sum(g["gain_eur"] for g in gains if g["gain_eur"] > 0)
        neg    = sum(g["gain_eur"] for g in gains if g["gain_eur"] < 0)
        net    = r2(pos + neg)
        tax    = sum(g["tax_amount"] for g in gains)

        print(f"Povzetek{label}:")
        print(f"  Skupni dobicki :  {pos:>12.2f} EUR")
        print(f"  Skupne izgube  :  {neg:>12.2f} EUR")
        print(f"  Neto           :  {net:>12.2f} EUR")
        print(f"  Ocena davka    :  {tax:>12.2f} EUR  (25 % / 20 % / 15 % / 10 % / 0 %)")

        print()
        print("Po vrednostnem papirju:")
        by_t = {}
        for g in gains:
            k = g["ticker"]
            if k not in by_t:
                by_t[k] = {"gain": 0.0, "tax": 0.0}
            by_t[k]["gain"] = r2(by_t[k]["gain"] + g["gain_eur"])
            by_t[k]["tax"]  = r2(by_t[k]["tax"]  + g["tax_amount"])
        for ticker, v in sorted(by_t.items(), key=lambda x: -x[1]["gain"]):
            sign = "+" if v["gain"] >= 0 else ""
            print(f"  {ticker:<12} {sign}{v['gain']:>10.2f} EUR   davek: {v['tax']:>8.2f} EUR")

    if unrealized:
        print()
        print(f"Odprte pozicije ({len(unrealized)} lotov):")
        by_t = {}
        for l in unrealized:
            k = l["ticker"]
            by_t.setdefault(k, 0.0)
            by_t[k] = r2(by_t[k] + l["qty_remaining"])
        for ticker, qty in sorted(by_t.items()):
            print(f"  {ticker:<12} {qty} enot (se ni prodano)")

    print()


# ================================================================
#  MAIN
# ================================================================

def main():
    ap = argparse.ArgumentParser(
        description="Davcni tracker -- FIFO kapitalski dobicki (ZDoh-2)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Primeri:
  python davek.py
  python davek.py --ibkr activity.csv --leto 2024
  python davek.py --transakcije moje.csv --xml --profil profil.json
  python davek.py --leto 2024 --izhod rezultati-2024
""",
    )
    ap.add_argument("--ibkr",         metavar="DATOTEKA.csv",
                    help="Uvozi iz IBKR Activity Statement CSV")
    ap.add_argument("--transakcije",  metavar="DATOTEKA.csv", default="transakcije.csv",
                    help="CSV s transakcijami (privzeto: transakcije.csv)")
    ap.add_argument("--leto",         type=int,
                    help="Filtriraj po letu (npr. 2024)")
    ap.add_argument("--xml",          action="store_true",
                    help="Izvozi eDavki XML (Doh-KDVP)")
    ap.add_argument("--profil",       metavar="profil.json",
                    help="JSON z osebnimi podatki za XML glavo")
    ap.add_argument("--izhod",        metavar="IME", default="davek",
                    help="Osnovno ime izhodnih datotek (privzeto: davek)")
    args = ap.parse_args()

    print("=" * 54)
    print("  Davcni tracker v1.0 -- engineeringinvestor.si/davek")
    print("=" * 54)

    # --- Branje ---
    trades   = []
    warnings = []

    if args.ibkr:
        p = Path(args.ibkr)
        if not p.exists():
            print(f"\nNAPAKA: '{args.ibkr}' ne obstaja.")
            sys.exit(1)
        print(f"\nBeram IBKR CSV: {args.ibkr}")
        trades, warnings = read_ibkr_csv(p)
        print(f"  Uvozeno: {len(trades)} transakcij")
        if warnings:
            print(f"  Opozorila ({len(warnings)}):")
            for w in warnings[:10]:
                print(f"    - {w}")
        missing_rate = sum(1 for t in trades if t.get("tecaj_eur", 0) == 0)
        if missing_rate:
            print(f"\n  POZOR: {missing_rate} transakcij nima tecaja EUR (IBKR ga ne izvozi).")
            print("  Odpri izhodni Excel in nastavi tecaje v stolpcu 'Tecaj EUR'.")
            print("  Tecaje najdes na: https://www.bsi.si/financni-podatki/tecajnica-bsi")
    else:
        p = Path(args.transakcije)
        if not p.exists():
            print(f"\nNAPAKA: '{args.transakcije}' ne obstaja.")
            print("Ustvari jo po vzorcu iz transakcije-vzorec.csv")
            sys.exit(1)
        print(f"\nBeram transakcije: {args.transakcije}")
        trades = read_transactions_csv(p)
        print(f"  Prebrano: {len(trades)} transakcij")

    if not trades:
        print("\nNi transakcij za obdelavo.")
        sys.exit(0)

    # --- FIFO ---
    print("\nIzracunam FIFO ...")
    realized, unrealized, errors = calc_fifo(trades)
    print(f"  Realizirani FIFO pari : {len(realized)}")
    print(f"  Odprte pozicije       : {len(unrealized)}")

    print_summary(realized, unrealized, errors, args.leto)

    # --- Excel ---
    suffix    = f"-{args.leto}" if args.leto else ""
    xlsx_file = f"{args.izhod}{suffix}.xlsx"
    print(f"Izvazam Excel: {xlsx_file}")
    ok = export_excel(trades, realized, unrealized, args.leto, xlsx_file)
    if ok:
        print(f"  Shranjeno: {xlsx_file}")

    # --- XML ---
    if args.xml:
        year = args.leto or datetime.now().year
        profile = {}
        if args.profil:
            try:
                with open(args.profil, encoding="utf-8") as f:
                    profile = json.load(f)
                print(f"\nBeram profil: {args.profil}")
            except Exception as e:
                print(f"\nOpozorilo: Ne morem prebrati profila ({e}).")
                print("XML bo generiran brez osebnih podatkov.")
        else:
            print("\nOpomba: Za XML z osebnimi podatki dodaj --profil profil.json")

        xml_file = f"Doh-KDVP-{year}.xml"
        print(f"Izvazam XML: {xml_file}")
        ok = export_edavki_xml(realized, profile, year, xml_file)
        if ok:
            print(f"  Shranjeno: {xml_file}")

    print("Koncano.\n")


if __name__ == "__main__":
    main()
